import os
import zipfile
import warnings
import openpyxl
import PySimpleGUI as sg


def consolidate(responses_folder, template_file, output_file):
    """UserWarnings are thrown for formatting weirdness with xlsx. Ignore them."""
    warnings.filterwarnings('ignore')
    template = openpyxl.load_workbook(template_file, data_only=True)
    smelter_lookup = template['Smelter Look-up']
    for smelter_row in smelter_lookup.rows:
        if metals.__contains__(smelter_row[0].value):
            if smelter_row[4].value is not None:
                approved_smelters.append(smelter_row[4].value)

    for file in os.listdir(responses_folder):
        if file.lower().endswith('.xlsx'):
            unapproved_rows = []
            try:
                xlsx_sheet = openpyxl.load_workbook(os.path.join(responses_folder, file), data_only=True)["Smelter List"]
                for r in xlsx_sheet.iter_rows(min_row=5, min_col=1):
                    cid = r[0].value
                    if cid is None:
                        cid = r[5].value
                    metal = r[1].value
                    if cid is None and r[3].value is None:
                        """Smelter will have one of these two values. Skip."""
                    elif cid.startswith('CID') and approved_smelters.__contains__(cid):
                        if metal == 'Gold':
                            au_list[cid] = r
                        elif metal == 'Tin':
                            sn_list[cid] = r
                        elif metal == 'Tungsten':
                            w_list[cid] = r
                        elif metal == 'Tantalum':
                            ta_list[cid] = r
                    else:
                        unapproved_rows.append(r)
            except zipfile.BadZipFile:
                print(file + ' did not having any smelters in Smelter List. Please verify!')
                """ignore"""

            if len(unapproved_rows):
                unapproved_smelters[file] = unapproved_rows

    all_smelters = list(au_list.values()) + list(sn_list.values()) + list(w_list.values()) + list(ta_list.values())
    spectra_list = template['Smelter List']
    row_count = 5
    for smelter in all_smelters:
        cell_count = 1
        for cell in smelter:
            spectra_list.cell(row=row_count, column=cell_count).value = cell.value
            cell_count += 1
        row_count += 1

    template.save(output_file+".xlsx")

    if len(unapproved_smelters):
        print('There are unapproved smelters included in this list. Please check unapproved_smelters.xlsx for details.')
        row_count = 1
        unapproved_file = openpyxl.Workbook()
        unapproved_sheet = unapproved_file.active
        for file, bad_smelters in unapproved_smelters.items():
            unapproved_sheet.cell(row=row_count, column=1, value=file)
            for bad_smelter in bad_smelters:
                cell_count = 2
                for cell in bad_smelter:
                    unapproved_sheet.cell(row=row_count, column=cell_count, value=cell.value)
                    cell_count += 1
                row_count += 1
        unapproved_file.save('unapproved_smelters.xlsx')


au_list = {}
sn_list = {}
w_list = {}
ta_list = {}
metals = {'Gold', 'Tin', 'Tungsten', 'Tantalum'}
approved_smelters = []
unapproved_smelters = {}

sg.theme('Reddit')
layout = [[sg.Text('Select responses folder: '), sg.Text(' '), sg.Input(key='folder'), sg.FolderBrowse()],
          [sg.Text('Select the CMRT template: '), sg.Input(key='template'), sg.FileBrowse()],
          [sg.Text('Save file as: '), sg.Text(' '*18), sg.Input(key='output'), sg.Text('.xlsx')],
          [sg.Button('Consolidate')]]
window = sg.Window('Consolidate CMRT responses', layout)

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED:
        break
    elif event == 'Consolidate':
        consolidate(values['folder'], values['template'], values['output'])
        break
